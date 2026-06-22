"""
Excel file import functionality for SHACL Creator
Handles conversion of Excel files to SHACL TTL format
"""

import io
import csv
from typing import List, Tuple

try:
    from ..exports import export_ttl_content
    from ..imports.csv_importer import csv_to_ttl
except ImportError:
    from exports import export_ttl_content
    from imports.csv_importer import csv_to_ttl


def get_excel_sheet_names(file_content: bytes) -> List[str]:
    """
    Extract sheet names from an Excel file
    
    Args:
        file_content: Binary content of Excel file
        
    Returns:
        List of sheet names in the workbook
        
    Raises:
        Exception: If file cannot be read as Excel format
    """
    import openpyxl
    
    wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    sheets = wb.sheetnames
    wb.close()
    return sheets


def import_excel_file(file_content: bytes, dataset_name: str, sheet_name: str = None, 
                      lang: str = 'de', editor=None) -> Tuple[bool, str]:
    """
    Import an Excel file and convert to SHACL TTL
    
    Args:
        file_content: Binary content of Excel file
        dataset_name: Name for the imported dataset
        sheet_name: Specific sheet to import (defaults to active sheet)
        lang: Language code for labels (de, en, fr, it)
        editor: FlaskSHACLGraphEditor instance for updating structure
        
    Returns:
        Tuple of (success: bool, message: str)
    """
    import openpyxl
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        
        # Select sheet
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        # Convert sheet to CSV string
        output = io.StringIO()
        writer = csv.writer(output)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(['' if v is None else str(v) for v in row])
        csv_data = output.getvalue()
        wb.close()
        
        # Convert CSV to TTL
        ttl = csv_to_ttl(csv_data, dataset_name, lang)
        
        if not ttl:
            return False, "Failed to convert Excel sheet to TTL"
        
        # If editor is provided, process the TTL import
        if editor:
            try:
                from .ttl_importer import process_csv_ttl_import
                process_csv_ttl_import(editor, ttl, f"{dataset_name}.xlsx", dataset_name)
            except Exception as e:
                # Continue with basic import even if advanced processing fails
                pass
        
        return True, "Excel file imported successfully"
        
    except Exception as e:
        return False, f"Failed to import Excel file: {str(e)}"
